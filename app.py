# -*- coding: utf-8 -*-
import streamlit as st
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
from matplotlib.ticker import FuncFormatter
from io import BytesIO
from datetime import datetime

# ======= Configurações globais =======
pd.options.mode.copy_on_write = True
st.set_page_config(page_title="Controle de Manutenção", layout="wide")

# ======= Formatter contábil (BR) para gráficos =======
def br_currency(x, pos=None):
    """Formata em contábil BR: R$ 1.234,56 e (R$ 1.234,56) para negativos."""
    try:
        val = float(x)
    except Exception:
        val = 0.0
    s = f"{abs(val):,.2f}"  # 1,234,567.89
    s = s.replace(",", "X").replace(".", ",").replace("X", ".")  # 1.234.567,89
    if val < 0:
        return f"(R$ {s})"
    return f"R$ {s}"

currency_fmt = FuncFormatter(br_currency)

# ======= Helpers de limpeza/normalização =======
def _make_unique_cols(cols):
    """
    - Normaliza nomes (remove \n, trim)
    - Substitui vazios/Unnamed por BLANK_<idx>
    - Garante unicidade com sufixo __1, __2, ...
    """
    out, seen = [], {}
    for i, c in enumerate(cols):
        name = "" if c is None else str(c)
        name = name.replace("\n", " ").strip()
        if name == "" or name.lower().startswith("unnamed"):
            name = f"BLANK_{i+1}"
        if name in seen:
            seen[name] += 1
            name = f"{name}__{seen[name]}"
        else:
            seen[name] = 0
        out.append(name)
    return out

def _normalize_cols(df: pd.DataFrame) -> pd.DataFrame:
    """
    Remove colunas 100% vazias (por índice, para não conflitar com nomes duplicados)
    e depois renomeia garantindo unicidade dos nomes.
    """
    df = df.copy()
    empty_idx = []
    for i in range(df.shape[1]):
        col_series = df.iloc[:, i]
        if not col_series.notna().any():
            empty_idx.append(i)
    if empty_idx:
        df.drop(df.columns[empty_idx], axis=1, inplace=True)
    df.columns = _make_unique_cols(df.columns)
    return df

def _to_number_br_series(s: pd.Series) -> pd.Series:
    """
    Converte série para numérico aceitando:
    - 1.234,56 (pt-BR) -> 1234.56
    - 1,234.56 (en-US) -> 1234.56
    - números já numéricos
    """
    def conv(x):
        if pd.isna(x):
            return np.nan
        if isinstance(x, (int, float, np.number)):
            return float(x)
        t = str(x).strip()
        # Caso típico BR com vírgula decimal (última vírgula depois do último ponto)
        if "," in t and (t.rfind(",") > t.rfind(".")):
            t = t.replace(".", "").replace(",", ".")
        try:
            return float(t)
        except Exception:
            return np.nan
    return s.apply(conv)

def _coerce_types_req(df: pd.DataFrame) -> pd.DataFrame:
    """
    Pós-edição: força tipos nas requisições para manter cálculos e gráficos estáveis.
    - VALOR -> numérico (aceita '1.234,56' e '1234.56')
    - Datas relevantes -> datetime
    - Recalcula 'MÊS' (YYYY-MM) a partir de 'MÊS COMPETÊNCIA' quando existir
    """
    d = df.copy()

    if "VALOR" in d.columns:
        d["VALOR"] = _to_number_br_series(d["VALOR"]).fillna(0.0)

    # Datas
    for col in [
        "MÊS COMPETÊNCIA","DATA DE CRIAÇÃO","Data Aprovação","DATA RECEBIMENTO",
        "DATA DO DOC","DATA DE ENTRADA","DATA DE LANÇAMENTO"
    ]:
        if col in d.columns:
            d[col] = pd.to_datetime(d[col], errors="coerce")

    # Campo MÊS (YYYY-MM)
    if "MÊS COMPETÊNCIA" in d.columns and pd.api.types.is_datetime64_any_dtype(d["MÊS COMPETÊNCIA"]):
        d["MÊS"] = d["MÊS COMPETÊNCIA"].dt.to_period("M").astype(str)
    else:
        if "MÊS COMPETÊNCIA" in d.columns:
            _m = pd.to_datetime(d["MÊS COMPETÊNCIA"], errors="coerce")
            d["MÊS"] = _m.dt.to_period("M").astype(str)
        else:
            if "MÊS" not in d.columns:
                d["MÊS"] = pd.NaT

    # Ajuste de 'CD ' -> 'CD'
    if "CD" not in d.columns and "CD " in d.columns:
        d["CD"] = d["CD "]

    return d

# ======= Carregamento OTIMIZADO do Excel =======
@st.cache_data(show_spinner=True)
def load_requisicoes_smart(xlsx_path: str, sheet_hint: str = "requis") -> pd.DataFrame:
    """
    Lê a planilha de Requisições em modo streaming (openpyxl read_only),
    parando após um bloco grande de linhas 100% vazias.
    """
    from openpyxl import load_workbook

    wb = load_workbook(xlsx_path, data_only=True, read_only=True)
    target_name = None
    for nm in wb.sheetnames:
        if sheet_hint.lower() in nm.lower():
            target_name = nm
            break
    ws = wb[target_name or wb.sheetnames[0]]

    rows = ws.iter_rows(values_only=True)
    header = None

    for r in rows:
        if r and any(r):
            header = [str(c).strip() if c is not None else "" for c in r]
            break
    if not header:
        return pd.DataFrame()

    data = []
    empty_streak = 0
    EMPTY_STOP = 400  # ajuste se necessário

    for r in rows:
        if not r or not any(r):
            empty_streak += 1
            if empty_streak >= EMPTY_STOP:
                break
            continue
        empty_streak = 0
        data.append(list(r))

    df = pd.DataFrame(data, columns=header)
    df = df.dropna(how="all")
    df = _normalize_cols(df)

    if "VALOR" in df.columns:
        df["VALOR"] = pd.to_numeric(df["VALOR"], errors="coerce").fillna(0.0)

    for col in [
        "MÊS COMPETÊNCIA","DATA DE CRIAÇÃO","Data Aprovação","DATA RECEBIMENTO",
        "DATA DO DOC","DATA DE ENTRADA","DATA DE LANÇAMENTO"
    ]:
        if col in df.columns:
            df[col] = pd.to_datetime(df[col], errors="coerce")

    if "MÊS COMPETÊNCIA" in df.columns and pd.api.types.is_datetime64_any_dtype(df["MÊS COMPETÊNCIA"]):
        df["MÊS"] = df["MÊS COMPETÊNCIA"].dt.to_period("M").astype(str)
    else:
        if "MÊS COMPETÊNCIA" in df.columns:
            _m = pd.to_datetime(df["MÊS COMPETÊNCIA"], errors="coerce")
            df["MÊS"] = _m.dt.to_period("M").astype(str)
        else:
            df["MÊS"] = pd.NaT

    if "CD" not in df.columns and "CD " in df.columns:
        df["CD"] = df["CD "]

    return df

@st.cache_data(show_spinner=True)
def load_budget(xlsx_path: str, sheet_hint: str = "budget") -> pd.DataFrame:
    xl = pd.ExcelFile(xlsx_path, engine="openpyxl")
    name = next((nm for nm in xl.sheet_names if sheet_hint.lower() in nm.lower()), xl.sheet_names[0])
    df = pd.read_excel(xl, sheet_name=name, dtype_backend="pyarrow")
    return _normalize_cols(df)

@st.cache_data(show_spinner=True)
def load_estornos(xlsx_path: str, sheet_hint: str = "estono") -> pd.DataFrame:
    try:
        xl = pd.ExcelFile(xlsx_path, engine="openpyxl")
        name = next((nm for nm in xl.sheet_names if sheet_hint.lower() in nm.lower()), None)
        if not name:
            return pd.DataFrame()
        df = pd.read_excel(xl, sheet_name=name, dtype_backend="pyarrow")
        return _normalize_cols(df)
    except Exception:
        return pd.DataFrame()

@st.cache_data(show_spinner=True)
def load_data():
    xlsx_path = "data/controle.xlsx"
    req_df = load_requisicoes_smart(xlsx_path, sheet_hint="requis")
    budget_df = load_budget(xlsx_path, sheet_hint="budget")
    estorno_df = load_estornos(xlsx_path, sheet_hint="estono")
    return {"req": req_df, "budget": budget_df, "estorno": estorno_df}

# ======= Carrega dados =======
data = load_data()
req_df = data["req"]
budget_df = data["budget"]
estorno_df = data["estorno"]

if req_df.empty:
    st.error("Não foi possível carregar a planilha de Requisições (aba 'Base_Requisicoes'). Verifique o arquivo em data/controle.xlsx.")
    st.stop()

# ======= Filtros =======
st.sidebar.title('Filtros')

cd_opts = sorted([x for x in req_df.get('CD', pd.Series(dtype=str)).dropna().astype(str).unique()])
cd_sel = st.sidebar.multiselect('CD (Centro de Distribuição)', cd_opts, default=cd_opts)

ano_opts = sorted([int(x) for x in pd.to_numeric(req_df.get('ANO', pd.Series()), errors='coerce').dropna().unique()])
ano_sel = st.sidebar.multiselect('Ano', ano_opts, default=ano_opts)

mes_opts = sorted(req_df['MÊS'].dropna().unique())
mes_sel = st.sidebar.multiselect('Mês competência', mes_opts, default=mes_opts)

grupo_opts = sorted(req_df.get('Grupo', pd.Series(dtype=str)).dropna().astype(str).unique())
grupo_sel = st.sidebar.multiselect('Grupo', grupo_opts, default=grupo_opts)

subgrupo_opts = sorted(req_df.get('SubGrupo', pd.Series(dtype=str)).dropna().astype(str).unique())
subgrupo_sel = st.sidebar.multiselect('Subgrupo', subgrupo_opts, default=subgrupo_opts)

status_opts = sorted(req_df.get('STATUS', pd.Series(dtype=str)).dropna().astype(str).unique())
status_sel = st.sidebar.multiselect('Status Requisição', status_opts, default=status_opts)

# Aplica filtros na base
f = (
    req_df
    .assign(
        CD=lambda d: d.get('CD', pd.Series(dtype=str)).astype(str),
        ANO=lambda d: pd.to_numeric(d.get('ANO', pd.Series()), errors='coerce'),
        Grupo=lambda d: d.get('Grupo', pd.Series(dtype=str)).astype(str),
        SubGrupo=lambda d: d.get('SubGrupo', pd.Series(dtype=str)).astype(str),
        STATUS=lambda d: d.get('STATUS', pd.Series(dtype=str)).astype(str),
    )
)
if cd_sel:
    f = f[f['CD'].isin(cd_sel)]
if ano_sel:
    f = f[f['ANO'].isin(ano_sel)]
if mes_sel:
    f = f[f['MÊS'].isin(mes_sel)]
if grupo_sel:
    f = f[f['Grupo'].isin(grupo_sel)]
if subgrupo_sel:
    f = f[f['SubGrupo'].isin(subgrupo_sel)]
if status_sel:
    f = f[f['STATUS'].isin(status_sel)]

# ======= Tabela EDITÁVEL =======
st.subheader('Requisições filtradas (editável)')
st.caption('As edições não sobrescrevem o Excel original. Baixe as edições para atualizar sua base quando quiser.')

edited_f = st.data_editor(
    f,
    num_rows="dynamic",
    use_container_width=True,
    height=420,
    key="edit_req"
)

# Força tipos e campos calculados após edição
f_used = _coerce_types_req(edited_f)

# ======= KPIs (baseados na tabela editada) =======
def format_br_number(val):
    try:
        return br_currency(float(val))
    except Exception:
        return "R$ 0,00"

c1, c2, c3, c4 = st.columns(4)
with c1:
    st.metric('Requisições (itens)', f_used.shape[0])
with c2:
    vtot = f_used['VALOR'].sum() if 'VALOR' in f_used.columns else 0.0
    st.metric('Valor total', format_br_number(vtot))
with c3:
    ticket = (vtot / f_used.shape[0]) if ('VALOR' in f_used.columns and f_used.shape[0]) else 0.0
    st.metric('Ticket médio', format_br_number(ticket))
with c4:
    aprovados = f_used[f_used.get('STATUS','')=='APROVADO']['VALOR'].sum() if 'STATUS' in f_used.columns and 'VALOR' in f_used.columns else 0.0
    st.metric('Aprovado', format_br_number(aprovados))

st.markdown('---')

# ======= Gráficos (com base em f_used) =======
g1, g2 = st.columns(2)

with g1:
    if 'MÊS' in f_used.columns and 'VALOR' in f_used.columns:
        by_mes = f_used.groupby('MÊS', dropna=True)['VALOR'].sum().sort_index()
        fig, ax = plt.subplots(figsize=(6,3))
        by_mes.plot(kind='bar', ax=ax, color='#1f77b4')
        ax.set_title('Despesas por mês (R$)')
        ax.set_xlabel('Mês')
        ax.set_ylabel('Valor (R$)')
        ax.yaxis.set_major_formatter(currency_fmt)
        st.pyplot(fig)

with g2:
    if 'Grupo' in f_used.columns and 'VALOR' in f_used.columns:
        by_grupo = f_used.groupby('Grupo')['VALOR'].sum().sort_values(ascending=False).head(10)
        fig2, ax2 = plt.subplots(figsize=(6,3))
        by_grupo.plot(kind='barh', ax=ax2, color='#2ca02c')
        ax2.set_title('Top 10 grupos por valor (R$)')
        ax2.set_xlabel('Valor (R$)')
        ax2.set_ylabel('Grupo')
        ax2.xaxis.set_major_formatter(currency_fmt)
        st.pyplot(fig2)

st.markdown('---')

# ======= BGT x REQ (a partir da tabela filtrada + editada) =======
# ======= BGT x REQ (a partir da tabela filtrada + editada) =======
if budget_df is not None and not budget_df.empty:
    st.subheader('Orçado x Requisitado (a partir da tabela acima)')

    if not f_used.empty and 'VALOR' in f_used.columns:
        # --- REQ por mês diretamente da tabela editável ---
        if 'MÊS' not in f_used.columns and 'MÊS COMPETÊNCIA' in f_used.columns:
            tmp_m = pd.to_datetime(f_used['MÊS COMPETÊNCIA'], errors='coerce')
            f_used = f_used.assign(MÊS=tmp_m.dt.to_period('M').astype(str))

        req_mes = (
            f_used.dropna(subset=['MÊS'])
                  .groupby('MÊS', dropna=False)['VALOR']
                  .sum()
                  .reset_index(name='REQ_VALOR')
        )

        # --- BGT por mês, filtrado conforme dimensões presentes na tabela ---
        budgetU = budget_df.copy()
        budgetU.columns = [str(c).strip().upper() for c in budgetU.columns]
        reqU = f_used.copy()
        reqU.columns = [str(c).strip().upper() for c in reqU.columns]

        # 1) Força NÚMERICO nas colunas de meses do Budget
        meses = ['JANEIRO','FEVEREIRO','MARÇO','ABRIL','MAIO','JUNHO',
                 'JULHO','AGOSTO','SETEMBRO','OUTUBRO','NOVEMBRO','DEZEMBRO']
        for m in meses:
            if m in budgetU.columns:
                budgetU[m] = _to_number_br_series(budgetU[m]).fillna(0.0)

        # 2) Filtro resiliente por dimensões (se zerar o conjunto, ignora o filtro)
        dim_pairs = [
            ('CD', 'CD'),
            ('GRUPO', 'GRUPO'),
            ('SUBGRUPO', 'SUBGRUPO'),
            ('CONTA', 'CONTA'),
            ('CENTRO DE CUSTO', 'CENTRO DE CUSTO'),
            ('CÓD.', 'CÓD. (CONTA+CENTRO)'),
        ]
        mask = pd.Series(True, index=budgetU.index)
        for bgt_col, req_col in dim_pairs:
            if bgt_col in budgetU.columns and req_col in reqU.columns:
                values = reqU[req_col].dropna().astype(str).unique()
                if len(values) > 0:
                    cand = mask & budgetU[bgt_col].astype(str).isin(values)
                    # só aplica se não “zera” tudo
                    if cand.any():
                        mask = cand
        budgetU = budgetU[mask]

        # 3) Derrete meses e cria coluna 'MÊS' (vetorizado, sem apply)
        present = [m for m in meses if m in budgetU.columns]
        if present and not budgetU.empty:
            bgt_long = budgetU.melt(
                id_vars=[c for c in budgetU.columns if c not in present],
                value_vars=present,
                var_name='MES_NOME', value_name='BGT_VALOR'
            )
            bgt_long['BGT_VALOR'] = _to_number_br_series(bgt_long['BGT_VALOR']).fillna(0.0)

            mapa_mes = {mes: i+1 for i, mes in enumerate(meses)}
            bgt_long['MES_NUM'] = bgt_long['MES_NOME'].map(mapa_mes)

            # ANO: tenta usar do Budget; se não houver, derive do filtro de Ano ou do ano atual
            if 'ANO' in bgt_long.columns:
                bgt_long['ANO'] = pd.to_numeric(bgt_long['ANO'], errors='coerce')
            else:
                bgt_long['ANO'] = pd.Series(np.nan, index=bgt_long.index)

            # Escopo de anos: 1) filtro "Ano", 2) anos de f_used/MÊS, 3) ano atual
            anos_scope = []
            if 'ano_sel' in locals() and ano_sel:
                anos_scope = sorted(set(int(a) for a in ano_sel))
            if not anos_scope and 'MÊS' in req_mes.columns and not req_mes.empty:
                anos_from_req = pd.to_numeric(req_mes['MÊS'].str.slice(0, 4), errors='coerce').dropna().astype(int).unique()
                anos_scope = sorted(set(anos_from_req.tolist()))
            if not anos_scope:
                anos_scope = [datetime.now().year]

            bgt_long.loc[bgt_long['ANO'].isna(), 'ANO'] = anos_scope[0]
            bgt_long['ANO'] = bgt_long['ANO'].astype(int)

            # Remove QUALQUER coluna "MÊS" preexistente (evita duplicatas)
            if (bgt_long.columns == 'MÊS').any():
                bgt_long = bgt_long.loc[:, bgt_long.columns != 'MÊS']

            bgt_long['MES_NUM'] = pd.to_numeric(bgt_long['MES_NUM'], errors='coerce').fillna(1).astype(int).clip(1, 12)
            ano_s = bgt_long['ANO'].astype(int).astype(str).str.zfill(4)
            mes_s = bgt_long['MES_NUM'].astype(int).astype(str).str.zfill(2)
            bgt_long['MÊS'] = ano_s + '-' + mes_s

            bgt_mes = (bgt_long.groupby('MÊS', as_index=False)['BGT_VALOR'].sum())
        else:
            bgt_mes = pd.DataFrame(columns=['MÊS','BGT_VALOR'])

        # 4) Calendário completo 01..12 para os ANOS em escopo
        if 'ano_sel' in locals() and ano_sel:
            anos_scope = sorted(set(int(a) for a in ano_sel))
        else:
            # tenta anos do REQ; senão do BGT; senão ano atual
            anos_req = pd.to_numeric(req_mes['MÊS'].str.slice(0, 4), errors='coerce').dropna().astype(int).unique() if not req_mes.empty else []
            anos_bgt = pd.to_numeric(bgt_mes['MÊS'].str.slice(0, 4), errors='coerce').dropna().astype(int).unique() if not bgt_mes.empty else []
            anos_scope = sorted(set(list(anos_req) + list(anos_bgt))) or [datetime.now().year]

        cal = pd.DataFrame({
            'MÊS': [f"{a:04d}-{m:02d}" for a in anos_scope for m in range(1, 13)]
        })

        # 5) Junta calendário + BGT + REQ, preenche vazios com 0
        comp = (cal
                .merge(bgt_mes, on='MÊS', how='left')
                .merge(req_mes, on='MÊS', how='left')
                .fillna(0.0)
                .sort_values('MÊS'))

        # 6) Gráfico com formatação contábil
        fig3, ax3 = plt.subplots(figsize=(8,3))
        ax3.plot(comp['MÊS'], comp['BGT_VALOR'], marker='o', label='Orçado (BGT)')
        ax3.plot(comp['MÊS'], comp['REQ_VALOR'], marker='o', label='Requisitado (REQ)')
        ax3.set_title('BGT x REQ por mês')
        ax3.set_xlabel('Mês')
        ax3.set_ylabel('Valor (R$)')
        ax3.yaxis.set_major_formatter(currency_fmt)
        ax3.legend()
        plt.xticks(rotation=45)
        st.pyplot(fig3)

        comp['Disponível'] = comp['BGT_VALOR'] - comp['REQ_VALOR']
        st.dataframe(comp, use_container_width=True)

# ======= Estornos Abertos =======
if estorno_df is not None and not estorno_df.empty:
    st.subheader('Estornos Abertos')
    st.dataframe(estorno_df, use_container_width=True, height=300)

# ======= Download (usando a base editada) =======
st.markdown('---')

def to_excel_bytes(df_dict):
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        for name, df in df_dict.items():
            df.to_excel(writer, sheet_name=name[:31], index=False)
    output.seek(0)
    return output

colA, colB, colC = st.columns(3)
with colA:
    if st.button('Baixar Requisições (filtradas e editadas)'):
        bio = to_excel_bytes({'Requisicoes_EditarPainel': f_used})
        st.download_button(
            'Download Requisicoes_Editadas.xlsx',
            data=bio,
            file_name='Requisicoes_editadas.xlsx',
            mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
with colB:
    if budget_df is not None and not budget_df.empty:
        if st.button('Baixar Budget (completo)'):
            bio2 = to_excel_bytes({'Budget': budget_df})
            st.download_button(
                'Download Budget.xlsx',
                data=bio2,
                file_name='Budget.xlsx',
                mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
with colC:
    st.caption('Baixe suas edições e, se quiser, atualize sua base original depois.')

st.caption('© 2026 - Painel criado para apoiar controle de requisições, NF e budget de Manutenção.')